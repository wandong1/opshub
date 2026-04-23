#!/usr/bin/env python3
# -*- coding: utf-8 -*-
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

# 创建工作簿
wb = Workbook()
ws = wb.active
ws.title = "SREHub巡检指标"

# 定义表头
headers = ["产品栏", "监控类型", "metric名称", "metric类型", "metric描述", "标签说明"]
ws.append(headers)

# 设置表头样式
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

# 定义所有指标数据
metrics_data = [
    # 通用核心指标
    ["连接平台（巡检&拨测平台）", "拨测", "srehub_inspect_task_exec_total", "Counter", "调度任务总执行次数",
     "task_id: 任务ID task_name: 任务名称 task_type: 任务类型 business_group: 业务分组 owner: 负责人 schedule_mode: 触发方式"],
    ["连接平台（巡检&拨测平台）", "拨测", "srehub_inspect_task_success_total", "Counter", "调度任务执行成功次数",
     "task_id: 任务ID task_name: 任务名称 task_type: 任务类型 business_group: 业务分组 owner: 负责人 schedule_mode: 触发方式"],
    ["连接平台（巡检&拨测平台）", "拨测", "srehub_inspect_task_fail_total", "Counter", "调度任务执行失败次数",
     "task_id: 任务ID task_name: 任务名称 task_type: 任务类型 business_group: 业务分组 owner: 负责人 schedule_mode: 触发方式"],
    ["连接平台（巡检&拨测平台）", "拨测", "srehub_inspect_task_retry_total", "Counter", "调度任务重试总次数",
     "task_id: 任务ID task_name: 任务名称 task_type: 任务类型 business_group: 业务分组 owner: 负责人 schedule_mode: 触发方式"],
    ["连接平台（巡检&拨测平台）", "拨测", "srehub_inspect_task_abort_total", "Counter", "调度任务主动终止/取消次数（手动stop）",
     "task_id: 任务ID task_name: 任务名称 task_type: 任务类型 business_group: 业务分组 owner: 负责人 schedule_mode: 触发方式"],
    ["连接平台（巡检&拨测平台）", "拨测", "srehub_inspect_task_exec_duration_seconds", "Gauge", "本次任务执行耗时（单位：秒）",
     "task_id: 任务ID task_name: 任务名称 task_type: 任务类型 business_group: 业务分组 owner: 负责人 schedule_mode: 触发方式"],
    ["连接平台（巡检&拨测平台）", "拨测", "srehub_inspect_task_availability", "Gauge", "任务可用性比率（成功次数/总次数，0~1）",
     "task_id: 任务ID task_name: 任务名称 task_type: 任务类型 business_group: 业务分组 owner: 负责人 schedule_mode: 触发方式"],
    ["连接平台（巡检&拨测平台）", "拨测", "srehub_inspect_task_availability_gauge", "Gauge", "本次执行结果（1=成功，0=失败）",
     "task_id: 任务ID task_name: 任务名称 task_type: 任务类型 business_group: 业务分组 owner: 负责人 schedule_mode: 触发方式"],
    # Ping 拨测指标
    ["连接平台（巡检&拨测平台）", "拨测", "srehub_inspect_ping_avg_rtt_seconds", "Gauge", "Ping平均往返时延RTT（单位：秒）",
     "task_id task_name task_type=ping business_group owner schedule_mode target probe_name"],
    ["连接平台（巡检&拨测平台）", "拨测", "srehub_inspect_ping_min_rtt_seconds", "Gauge", "Ping最小时延（单位：秒）",
     "task_id task_name task_type=ping business_group owner schedule_mode target probe_name"],
    ["连接平台（巡检&拨测平台）", "拨测", "srehub_inspect_ping_max_rtt_seconds", "Gauge", "Ping最大时延（单位：秒）",
     "task_id task_name task_type=ping business_group owner schedule_mode target probe_name"],
    ["连接平台（巡检&拨测平台）", "拨测", "srehub_inspect_ping_jitter_seconds", "Gauge", "Ping时延抖动/RTT标准差（单位：秒）",
     "task_id task_name task_type=ping business_group owner schedule_mode target probe_name"],
    ["连接平台（巡检&拨测平台）", "拨测", "srehub_inspect_ping_loss_ratio", "Gauge", "Ping丢包率（0~1）",
     "task_id task_name task_type=ping business_group owner schedule_mode target probe_name"],
    ["连接平台（巡检&拨测平台）", "拨测", "srehub_inspect_ping_packet_send_total", "Counter", "Ping发送数据包总数（累计）",
     "task_id task_name task_type=ping business_group owner schedule_mode target probe_name"],
    ["连接平台（巡检&拨测平台）", "拨测", "srehub_inspect_ping_packet_recv_total", "Counter", "Ping接收数据包总数（累计）",
     "task_id task_name task_type=ping business_group owner schedule_mode target probe_name"],
    
    # TCP 拨测指标
    ["连接平台（巡检&拨测平台）", "拨测", "srehub_inspect_tcp_connect_success_total", "Counter", "TCP建连成功次数（累计）",
     "task_id task_name task_type=tcp business_group owner schedule_mode target probe_name"],
    ["连接平台（巡检&拨测平台）", "拨测", "srehub_inspect_tcp_connect_fail_total", "Counter", "TCP建连失败次数（累计）",
     "task_id task_name task_type=tcp business_group owner schedule_mode target probe_name"],
    ["连接平台（巡检&拨测平台）", "拨测", "srehub_inspect_tcp_connect_duration_seconds", "Gauge", "TCP建连耗时（单位：秒）",
     "task_id task_name task_type=tcp business_group owner schedule_mode target probe_name"],
    ["连接平台（巡检&拨测平台）", "拨测", "srehub_inspect_tcp_port_reachable", "Gauge", "TCP端口是否可达（1=可达，0=不可达）",
     "task_id task_name task_type=tcp business_group owner schedule_mode target probe_name"],
    
    # UDP 拨测指标
    ["连接平台（巡检&拨测平台）", "拨测", "srehub_inspect_udp_send_total", "Counter", "UDP报文发送总数（累计）",
     "task_id task_name task_type=udp business_group owner schedule_mode target probe_name"],
    ["连接平台（巡检&拨测平台）", "拨测", "srehub_inspect_udp_recv_total", "Counter", "UDP报文接收总数（累计）",
     "task_id task_name task_type=udp business_group owner schedule_mode target probe_name"],
    ["连接平台（巡检&拨测平台）", "拨测", "srehub_inspect_udp_loss_total", "Counter", "UDP报文丢失总数（累计）",
     "task_id task_name task_type=udp business_group owner schedule_mode target probe_name"],
    ["连接平台（巡检&拨测平台）", "拨测", "srehub_inspect_udp_transfer_delay_seconds", "Gauge", "UDP报文传输时延（写+读时延之和，单位：秒）",
     "task_id task_name task_type=udp business_group owner schedule_mode target probe_name"],
    # HTTP/HTTPS 拨测指标
    ["连接平台（巡检&拨测平台）", "拨测", "srehub_inspect_http_response_duration_seconds", "Gauge", "HTTP请求总响应耗时（建连→完整接收，单位：秒）",
     "task_id task_name task_type=http/https business_group owner schedule_mode target probe_name http_method http_path"],
    ["连接平台（巡检&拨测平台）", "拨测", "srehub_inspect_http_dns_duration_seconds", "Gauge", "HTTP DNS解析耗时（单位：秒）",
     "task_id task_name task_type=http/https business_group owner schedule_mode target probe_name http_method http_path"],
    ["连接平台（巡检&拨测平台）", "拨测", "srehub_inspect_http_tls_duration_seconds", "Gauge", "HTTP TLS握手耗时（单位：秒）",
     "task_id task_name task_type=http/https business_group owner schedule_mode target probe_name http_method http_path"],
    ["连接平台（巡检&拨测平台）", "拨测", "srehub_inspect_http_first_byte_seconds", "Gauge", "HTTP首字节响应时间TTFB（单位：秒）",
     "task_id task_name task_type=http/https business_group owner schedule_mode target probe_name http_method http_path"],
    ["连接平台（巡检&拨测平台）", "拨测", "srehub_inspect_http_assertion_result", "Gauge", "HTTP断言结果（1=通过，0=失败）",
     "task_id task_name task_type=http/https business_group owner schedule_mode target probe_name http_method http_path"],
    ["连接平台（巡检&拨测平台）", "拨测", "srehub_inspect_http_status_code_total", "Counter", "HTTP响应码分布计数（累计）",
     "task_id task_name task_type=http/https business_group owner schedule_mode status_code"],
    ["连接平台（巡检&拨测平台）", "拨测", "srehub_inspect_https_cert_valid_days", "Gauge", "HTTPS证书剩余有效天数（单位：天）",
     "task_id task_name task_type=https business_group owner schedule_mode target probe_name"],
    
    # WebSocket/WSS 拨测指标
    ["连接平台（巡检&拨测平台）", "拨测", "srehub_inspect_ws_connection_established", "Gauge", "WebSocket连接是否成功建立（1=成功，0=失败）",
     "task_id task_name task_type=websocket business_group owner schedule_mode target probe_name"],
    ["连接平台（巡检&拨测平台）", "拨测", "srehub_inspect_ws_handshake_duration_seconds", "Gauge", "WebSocket握手建连耗时（单位：秒）",
     "task_id task_name task_type=websocket business_group owner schedule_mode target probe_name"],
    ["连接平台（巡检&拨测平台）", "拨测", "srehub_inspect_ws_handshake_success_total", "Counter", "WebSocket握手成功次数（累计）",
     "task_id task_name task_type=websocket business_group owner schedule_mode target probe_name"],
    ["连接平台（巡检&拨测平台）", "拨测", "srehub_inspect_ws_disconnect_total", "Counter", "WebSocket异常断开连接次数（累计）",
     "task_id task_name task_type=websocket business_group owner schedule_mode target probe_name"],
    # 业务编排（Workflow）指标
    ["连接平台（巡检&拨测平台）", "拨测", "srehub_inspect_flow_step_exec_total", "Counter", "编排步骤执行次数（累计）",
     "task_id task_name task_type=probe_flow business_group owner schedule_mode flow_id step_id step_name"],
    ["连接平台（巡检&拨测平台）", "拨测", "srehub_inspect_flow_step_fail_total", "Counter", "编排步骤失败次数（累计）",
     "task_id task_name task_type=probe_flow business_group owner schedule_mode flow_id step_id step_name"],
    ["连接平台（巡检&拨测平台）", "拨测", "srehub_inspect_flow_step_status", "Gauge", "编排步骤执行状态（1=成功，0=失败，2=跳过）",
     "task_id task_name task_type=probe_flow business_group owner schedule_mode flow_id step_id step_name"],
    ["连接平台（巡检&拨测平台）", "拨测", "srehub_inspect_flow_step_exec_duration", "Gauge", "编排步骤执行耗时（单位：秒）",
     "task_id task_name task_type=probe_flow business_group owner schedule_mode flow_id step_id step_name"],
    ["连接平台（巡检&拨测平台）", "拨测", "srehub_inspect_flow_step_assert_result", "Gauge", "编排步骤断言结果（1=通过，0=失败，-1=无断言）",
     "task_id task_name task_type=probe_flow business_group owner schedule_mode flow_id step_id step_name"],
    
    # 智能巡检指标
    ["连接平台（巡检&拨测平台）", "巡检", "srehub_inspect_check_pass_total", "Counter", "巡检项合规（通过）次数（累计）",
     "task_id task_name task_type=inspect business_group owner schedule_mode check_group check_item check_level host_id"],
    ["连接平台（巡检&拨测平台）", "巡检", "srehub_inspect_check_fail_total", "Counter", "巡检项不合规（不通过）次数（累计）",
     "task_id task_name task_type=inspect business_group owner schedule_mode check_group check_item check_level host_id"],
    ["连接平台（巡检&拨测平台）", "巡检", "srehub_inspect_check_abnormal_total", "Counter", "巡检异常数（累计）",
     "task_id task_name task_type=inspect business_group owner schedule_mode check_group check_item check_level host_id"],
    ["连接平台（巡检&拨测平台）", "巡检", "srehub_inspect_check_status", "Gauge", "巡检项是否通过（1=通过，0=不通过）",
     "task_id task_name task_type=inspect business_group owner schedule_mode check_group check_item check_level host_id"],
    ["连接平台（巡检&拨测平台）", "巡检", "srehub_inspect_check_duration_seconds", "Gauge", "巡检项执行耗时（单位：秒）",
     "task_id task_name task_type=inspect business_group owner schedule_mode check_group check_item check_level host_id"],
    ["连接平台（巡检&拨测平台）", "巡检", "srehub_inspect_check_assertion_result", "Gauge", "巡检断言结果（1=通过，0=失败）",
     "task_id task_name task_type=inspect business_group owner schedule_mode check_group check_item check_level host_id"],
]

# 添加数据到工作表
for row_data in metrics_data:
    ws.append(row_data)

# 设置列宽
ws.column_dimensions['A'].width = 25
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 45
ws.column_dimensions['D'].width = 12
ws.column_dimensions['E'].width = 50
ws.column_dimensions['F'].width = 80

# 设置数据行样式
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    for cell in row:
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

# 保存文件
output_file = "/Users/Zhuanz/golang_project/src/opshub/SREHub巡检指标.xlsx"
wb.save(output_file)
print(f"Excel 文件已生成: {output_file}")
print(f"共包含 {len(metrics_data)} 个指标")
